"""
匯入言語知識題庫。

讀取 data/raw/ 底下的 Excel 檔案，檔案可能有多個工作表，每個工作表都是題目（依序合併
匯入，question_number 跨所有工作表連續編號）。每個工作表的選項數可能不同（3 個 A/B/C
或 4 個 A/B/C/D，程式會依實際填寫的欄位動態判斷），題目裡的挖空標記也可能不同（「＿＿」
或「（　　）」皆有，程式會自動用正則從題目文字裡抓出實際用的是哪一種；少數題目本身沒有
挖空、是直接問文法用法的題型，這種就沒有 blank_marker，維持 None）。

用法：
    # 先跑小批次（前 5 題，跨工作表累計）確認流程沒問題
    python scripts/import_language_knowledge_questions.py \\
        --file "data/raw/檔名.xlsx" --exam-scope "範圍名稱" --limit 5

    # 確認沒問題後，正式全量匯入
    python scripts/import_language_knowledge_questions.py \\
        --file "data/raw/檔名.xlsx" --exam-scope "範圍名稱"

重複使用注意事項與 import_proverb_questions.py／import_vocab_questions.py 相同：全量匯入、
換範圍要換新的 --exam-scope、修改舊資料前要先確認 attempts_log 有沒有參照。
"""

import argparse
import re
import sys
from pathlib import Path
from typing import Any, Optional

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402

from app.db.client import supabase  # noqa: E402

OPTION_LETTERS = ["A", "B", "C", "D"]
BLANK_PATTERN = re.compile(r"＿+|（\s*）")


def _clean(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _detect_blank_marker(context_sentence: str) -> Optional[str]:
    match = BLANK_PATTERN.search(context_sentence)
    return match.group() if match else None


def _read_sheet(ws) -> list[dict[str, Any]]:
    header = [c.value for c in ws[1]]
    option_cols = [i for i, h in enumerate(header, start=1) if h and str(h).startswith("選項")]
    correct_col = header.index("正確答案") + 1
    explanation_col = header.index("解析") + 1 if "解析" in header else None

    rows = []
    for r in range(2, ws.max_row + 1):
        context_sentence = ws.cell(row=r, column=1).value
        if context_sentence is None or _clean(context_sentence) == "":
            break
        context_sentence = _clean(context_sentence)

        options = []
        for i, col in enumerate(option_cols):
            text = ws.cell(row=r, column=col).value
            if text is None or _clean(text) == "":
                continue
            options.append({"id": OPTION_LETTERS[i].lower(), "text": _clean(text)})

        correct_answer = _clean(ws.cell(row=r, column=correct_col).value).upper()
        explanation = _clean(ws.cell(row=r, column=explanation_col).value) if explanation_col else ""

        rows.append(
            {
                "context_sentence": context_sentence,
                "blank_marker": _detect_blank_marker(context_sentence),
                "options": options,
                "correct_option": correct_answer.lower(),
                "explanation_rule": explanation,
            }
        )
    return rows


def build_rows(
    entries: list[dict[str, Any]], exam_scope: str, start_number: int
) -> tuple[list[dict[str, Any]], int]:
    rows = []
    n = start_number
    for entry in entries:
        rows.append(
            {
                "mode": "language_knowledge",
                "exam_scope": exam_scope,
                "question_number": n,
                "stage": None,
                "context_sentence": entry["context_sentence"],
                "blank_marker": entry["blank_marker"],
                "options": entry["options"],
                "correct_option": entry["correct_option"],
                "explanation_rule": entry["explanation_rule"],
            }
        )
        n += 1
    return rows, n


def main() -> None:
    parser = argparse.ArgumentParser(description="匯入言語知識題庫（多工作表）")
    parser.add_argument("--file", required=True, help="Excel 檔案路徑（相對於專案根目錄）")
    parser.add_argument("--exam-scope", required=True, help="這批題目要匯入的 exam_scope 標籤")
    parser.add_argument("--limit", type=int, default=None, help="只匯入前 N 題（測試用，跨工作表累計）")
    args = parser.parse_args()

    existing = (
        supabase.table("questions")
        .select("id")
        .eq("mode", "language_knowledge")
        .eq("exam_scope", args.exam_scope)
        .execute()
    )
    if existing.data:
        print(f'錯誤：exam_scope="{args.exam_scope}" 底下已經有 {len(existing.data)} 筆 language_knowledge 題目，中止匯入。')
        print("如果是要修正/更新這個範圍的題庫，請先手動確認並清除舊資料（注意 attempts_log 等表格的參照）。")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.file, data_only=True)

    all_rows: list[dict[str, Any]] = []
    next_number = 1
    for sheet_name in wb.sheetnames:
        entries = _read_sheet(wb[sheet_name])
        sheet_rows, next_number = build_rows(entries, args.exam_scope, next_number)
        all_rows.extend(sheet_rows)
        print(f"  {sheet_name}: {len(sheet_rows)} 題")

    if args.limit:
        all_rows = all_rows[: args.limit]

    for row in all_rows:
        supabase.table("questions").insert(row).execute()

    print(f'已匯入 {len(all_rows)} 題，exam_scope="{args.exam_scope}"')


if __name__ == "__main__":
    main()
