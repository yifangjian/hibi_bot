"""
匯入諺語題庫。

讀取 data/raw/ 底下的 Excel 檔案（意味選択／文脈穴埋め／読み方三個工作表，依列順序
一一對應同一句諺語，第 2 列開始，跳過標題列），寫入 questions 表：同一句諺語寫入
3 筆，question_number 從 1 開始流水編號、三筆共用同一個號碼，用 stage 區分是哪一種
（semantic_choice / situational_choice / reading_input）。practice 出題時，
semantic_choice 與 situational_choice 兩種變體會隨機擇一當第一階段題目（見
app/services/question_picker.py 的 get_scope_candidates），reading_input 固定
接在第二階段。

重複使用注意事項（重要）：
- 這支腳本是「全量匯入」，不是增量／upsert。執行前會先檢查指定的 exam_scope 底下
  是否已經有 proverb 題目，如果有就直接中止，不會自動覆蓋或疊加，避免不小心把同一批
  題目匯入兩次造成重複。
- 如果是「全新的考試範圍」（例如下一次段考、小考），換一個新的 --exam-scope 字串即可，
  不會跟舊範圍衝突——question_number 只在同一個 (mode, exam_scope) 內唯一，不同範圍
  可以各自從 1 開始編號。
- 如果要「修正／更新」某個 exam_scope 已匯入的題庫內容（例如改了幾題的解析），需要
  先手動決定是否要刪除該 exam_scope 下的舊資料。務必先確認 attempts_log／
  wrong_question_state 等表格有沒有已經參照這些題目的作答紀錄——如果有（代表已經
  進入正式資料蒐集期），刪除前必須先跟研究者確認過，不能自動執行。這支腳本本身不會
  做任何刪除，需要的話請手動下 SQL。

用法：
    # 先跑小批次（前 5 句）確認流程沒問題
    python scripts/import_proverb_questions.py \\
        --file "data/raw/高日暑修_諺語題庫_期中100句.xlsx" \\
        --exam-scope "高日暑修班期中考" --limit 5

    # 確認沒問題後，正式全量匯入
    python scripts/import_proverb_questions.py \\
        --file "data/raw/高日暑修_諺語題庫_期中100句.xlsx" \\
        --exam-scope "高日暑修班期中考"
"""

import argparse
import sys
from pathlib import Path
from typing import Any, Optional

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402

from app.db.client import supabase  # noqa: E402

SHEET_SEMANTIC = "意味選択"
SHEET_SITUATIONAL = "文脈穴埋め"
SHEET_READING = "読み方"

OPTION_LETTERS = ["A", "B", "C", "D"]
BLANK_MARKER = "___"


def _clean(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _read_choice_sheet(ws) -> list[dict[str, Any]]:
    """讀取 意味選択／文脈穴埋め 這種「題目＋A/B/C/D＋正確答案＋解析」的工作表。"""
    rows = []
    for r in range(2, ws.max_row + 1):
        context_sentence = ws.cell(row=r, column=1).value
        if context_sentence is None or _clean(context_sentence) == "":
            break
        options = [
            {"id": letter.lower(), "text": _clean(ws.cell(row=r, column=2 + i).value)}
            for i, letter in enumerate(OPTION_LETTERS)
        ]
        correct_answer = _clean(ws.cell(row=r, column=6).value).upper()
        explanation = _clean(ws.cell(row=r, column=7).value)
        rows.append(
            {
                "context_sentence": _clean(context_sentence),
                "options": options,
                "correct_option": correct_answer.lower(),
                "explanation_rule": explanation,
            }
        )
    return rows


def _read_reading_sheet(ws) -> list[dict[str, Any]]:
    """讀取 読み方 工作表（題目＋正確答案，無選項、無解析）。"""
    rows = []
    for r in range(2, ws.max_row + 1):
        context_sentence = ws.cell(row=r, column=1).value
        if context_sentence is None or _clean(context_sentence) == "":
            break
        rows.append({"correct_option": _clean(ws.cell(row=r, column=2).value)})
    return rows


def build_rows(
    semantic: list[dict[str, Any]],
    situational: list[dict[str, Any]],
    reading: list[dict[str, Any]],
    exam_scope: str,
) -> list[dict[str, Any]]:
    if not (len(semantic) == len(situational) == len(reading)):
        raise ValueError(
            f"三個工作表列數不一致，請確認檔案內容：意味選択={len(semantic)} "
            f"文脈穴埋め={len(situational)} 読み方={len(reading)}"
        )

    rows = []
    for i, (sem, sit, yom) in enumerate(zip(semantic, situational, reading), start=1):
        rows.append(
            {
                "mode": "proverb",
                "exam_scope": exam_scope,
                "question_number": i,
                "stage": "semantic_choice",
                "context_sentence": sem["context_sentence"],
                "blank_marker": None,
                "options": sem["options"],
                "correct_option": sem["correct_option"],
                "explanation_rule": sem["explanation_rule"],
            }
        )
        situational_blank_marker: Optional[str] = BLANK_MARKER if BLANK_MARKER in sit["context_sentence"] else None
        rows.append(
            {
                "mode": "proverb",
                "exam_scope": exam_scope,
                "question_number": i,
                "stage": "situational_choice",
                "context_sentence": sit["context_sentence"],
                "blank_marker": situational_blank_marker,
                "options": sit["options"],
                "correct_option": sit["correct_option"],
                "explanation_rule": sit["explanation_rule"],
            }
        )
        rows.append(
            {
                "mode": "proverb",
                "exam_scope": exam_scope,
                "question_number": i,
                "stage": "reading_input",
                "context_sentence": None,
                "blank_marker": None,
                "options": None,
                "correct_option": yom["correct_option"],
                "explanation_rule": None,
            }
        )
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description="匯入諺語題庫（意味選択／文脈穴埋め／読み方）")
    parser.add_argument("--file", required=True, help="Excel 檔案路徑（相對於專案根目錄）")
    parser.add_argument("--exam-scope", required=True, help="這批題目要匯入的 exam_scope 標籤")
    parser.add_argument("--limit", type=int, default=None, help="只匯入前 N 句（測試用）")
    args = parser.parse_args()

    existing = (
        supabase.table("questions").select("id").eq("mode", "proverb").eq("exam_scope", args.exam_scope).execute()
    )
    if existing.data:
        print(f'錯誤：exam_scope="{args.exam_scope}" 底下已經有 {len(existing.data)} 筆 proverb 題目，中止匯入。')
        print("如果是要修正/更新這個範圍的題庫，請先手動確認並清除舊資料（注意 attempts_log 等表格的參照）。")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.file, data_only=True)
    semantic = _read_choice_sheet(wb[SHEET_SEMANTIC])
    situational = _read_choice_sheet(wb[SHEET_SITUATIONAL])
    reading = _read_reading_sheet(wb[SHEET_READING])

    rows = build_rows(semantic, situational, reading, args.exam_scope)

    if args.limit:
        rows = rows[: args.limit * 3]

    for row in rows:
        supabase.table("questions").insert(row).execute()

    sentence_count = len(rows) // 3
    print(f'已匯入 {sentence_count} 句諺語（共 {len(rows)} 筆），exam_scope="{args.exam_scope}"')


if __name__ == "__main__":
    main()
