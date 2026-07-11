"""
匯入單字讀音題庫。

讀取 data/raw/ 底下的 Excel 檔案（單一工作表：題目／選項A-D／正確答案，第 2 列開始，
跳過標題列），寫入 questions 表。這批題目是「詞彙讀音測驗」形式（題目欄位是單一詞彙，
不是情境例句），沒有解析欄位——単語模式只考讀音本身，答對答錯沒有需要 AI 說明的細膩
語感，所以答題後直接顯示正確讀音，不呼叫 OpenAI（見 app/services/menu_actions.py 的
_build_feedback_text）。

question_number 從 1 開始流水編號，只在同一個 (mode, exam_scope) 內唯一（見
app/db/schema.sql 的 unique_question_number_per_scope_stage）。

重複使用注意事項（跟 import_proverb_questions.py 相同）：
- 這是「全量匯入」腳本，不是增量／upsert。執行前會先檢查指定的 exam_scope 底下是否
  已經有 vocab 題目，如果有就直接中止，不會自動覆蓋或疊加。
- 換到全新的考試範圍：直接換一個新的 --exam-scope 字串即可。匯入後記得手動更新
  active_exam_scope 切到新範圍。
- 要修正/更新已匯入的內容：這支腳本不會做任何刪除，需要手動決定是否清除舊資料，
  動手前務必先查 attempts_log/wrong_question_state 有沒有已經參照這些題目的作答紀錄。

用法：
    # 先跑小批次（前 5 個詞）確認流程沒問題
    python scripts/import_vocab_questions.py \\
        --file "data/raw/日語讀音測驗_270題.xlsx" \\
        --exam-scope "高日暑修班期中考" --limit 5

    # 確認沒問題後，正式全量匯入
    python scripts/import_vocab_questions.py \\
        --file "data/raw/日語讀音測驗_270題.xlsx" \\
        --exam-scope "高日暑修班期中考"
"""

import argparse
import sys
from pathlib import Path
from typing import Any

sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

import openpyxl  # noqa: E402

from app.db.client import supabase  # noqa: E402

SHEET_NAME = "読み方クイズ"
OPTION_LETTERS = ["A", "B", "C", "D"]


def _clean(value: Any) -> str:
    return str(value).strip() if value is not None else ""


def _read_rows(ws) -> list[dict[str, Any]]:
    rows = []
    for r in range(2, ws.max_row + 1):
        word = ws.cell(row=r, column=1).value
        if word is None or _clean(word) == "":
            break
        options = [
            {"id": letter.lower(), "text": _clean(ws.cell(row=r, column=2 + i).value)}
            for i, letter in enumerate(OPTION_LETTERS)
        ]
        correct_answer = _clean(ws.cell(row=r, column=6).value).upper()
        rows.append({"word": _clean(word), "options": options, "correct_option": correct_answer.lower()})
    return rows


def build_rows(entries: list[dict[str, Any]], exam_scope: str) -> list[dict[str, Any]]:
    return [
        {
            "mode": "vocab",
            "exam_scope": exam_scope,
            "question_number": i,
            "stage": None,
            "context_sentence": entry["word"],
            "blank_marker": None,
            "options": entry["options"],
            "correct_option": entry["correct_option"],
            "explanation_rule": None,
        }
        for i, entry in enumerate(entries, start=1)
    ]


def main() -> None:
    parser = argparse.ArgumentParser(description="匯入單字讀音題庫（読み方クイズ）")
    parser.add_argument("--file", required=True, help="Excel 檔案路徑（相對於專案根目錄）")
    parser.add_argument("--exam-scope", required=True, help="這批題目要匯入的 exam_scope 標籤")
    parser.add_argument("--limit", type=int, default=None, help="只匯入前 N 個詞（測試用）")
    args = parser.parse_args()

    existing = (
        supabase.table("questions").select("id").eq("mode", "vocab").eq("exam_scope", args.exam_scope).execute()
    )
    if existing.data:
        print(f'錯誤：exam_scope="{args.exam_scope}" 底下已經有 {len(existing.data)} 筆 vocab 題目，中止匯入。')
        print("如果是要修正/更新這個範圍的題庫，請先手動確認並清除舊資料（注意 attempts_log 等表格的參照）。")
        sys.exit(1)

    wb = openpyxl.load_workbook(args.file, data_only=True)
    entries = _read_rows(wb[SHEET_NAME])

    rows = build_rows(entries, args.exam_scope)
    if args.limit:
        rows = rows[: args.limit]

    for row in rows:
        supabase.table("questions").insert(row).execute()

    print(f'已匯入 {len(rows)} 個單字，exam_scope="{args.exam_scope}"')


if __name__ == "__main__":
    main()
